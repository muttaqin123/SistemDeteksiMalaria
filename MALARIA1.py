from PIL import Image as imageee, ImageTk
from matplotlib import pyplot as plt
from time import sleep
from picamera import PiCamera,Color
from tkinter import messagebox
import numpy as np, glob, time,os, cv2, tkinter, pandas as pd
from tkinter import *
import tkinter.filedialog
from tkinter import ttk
from tkinter.ttk import Progressbar
from pandas import ExcelFile
from pandas import ExcelWriter
from picamera.renderers import PiRenderer
from library.StepperMotor import Lengan
from pynput import keyboard
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import tensorflow as tf
from tensorflow import keras
from keras.preprocessing import image
# ============================ MAIN CODE ============================

w = Tk()
a = '#178a4f'
btn = imageee.open("/home/pi/main-programaqin/Foto/nnew9.jpg")
w.btnn = ImageTk.PhotoImage(btn)
width_of_window = 1024
height_of_window = 600
screen_width = w.winfo_screenwidth()
screen_height = w.winfo_screenheight()
x_coordinate = (screen_width/2)-(width_of_window/2)
y_coordinate = (screen_height/2)-(height_of_window/2)
w.geometry("%dx%d+%d+%d" %
           (width_of_window, height_of_window, x_coordinate, y_coordinate))

w.attributes('-fullscreen',True)

s = ttk.Style()
s.theme_use('clam')
s.configure("red.Horizontal.TProgressbar",
            foreground='red', background='#4f4f4f')
progress = Progressbar(w, style="red.Horizontal.TProgressbar",
                       orient=HORIZONTAL, length=1300, mode='determinate',)

def new_win():
    ########Frame Configuration############
    q = Tk()
    q.title('SiDeMar Malaria')
    q.geometry('1024x600')
    q.config(bg='#178a4f')
    width_of_window = 1024
    height_of_window = 600
    screen_width = q.winfo_screenwidth()
    screen_height = q.winfo_screenheight()
    x_coordinate = (screen_width/2)-(width_of_window/2)
    y_coordinate = (screen_height/2)-(height_of_window/2)
    q.geometry("%dx%d+%d+%d" %
               (width_of_window, height_of_window, x_coordinate, y_coordinate))
    q.attributes('-fullscreen',True)
    
    b0 = Label(q,text='Sistem Deteksi',border=0, fg='white', bg=a, anchor='center')
    lst0 = ('Calibri (Body)', 32, 'bold')
    b0.config(font=lst0)
    b0.place(x=340, y=80)
    
    b00 = Label(q,text='Penyakit Malaria',border=0, fg='white', bg=a, anchor='center')
    b00.config(font=lst0)
    b00.place(x=320, y=130)
    
        ###### Button Main Windows ######
    b2 = Button(q, width=20, height=3, text='Open Camera',command=kamera,border=0, fg=a, bg='white')
    b2.place(x=660, y=450)

    b1 = Button(q, width=20, height=3, text='Open Photo Pasien',command=openimage,border=0, fg=a, bg='white')
    b1.place(x=850, y=450)

    b3 = Button(q, width=20, height=3, text='Deteksi Malaria',command=proses,border=0, fg=a, bg='white')
    b3.place(x=660, y=520)

    b4 = Button(q, width=20, height=3, text='Exit',command=q.destroy,border=0,fg=a, bg='white')
    b4.place(x=850, y=520)

    b5 = Button(q,width=20,height=3,text='Get Report',command=report,border=0,fg=a,bg='white')
    b5.place(x=470,y=520)

    b6 = Button(q,width=20,height=3,text='Help',command =help, border=0,fg=a,bg='white')
    b6.place(x=10,y=520)

    q.mainloop() 
         
#########Progress Bar############
def bar():

    l4 = Label(w, text='Loading...', fg='white', bg=a)
    lst4 = ('Calibri (Body)', 16)
    l4.config(font=lst4)
    l4.place(x=470, y=470)

    import time
    r = 0
    for i in range(50):
        progress['value'] = r
        w.update_idletasks()
        time.sleep(0.03)
        r = r+1

    w.destroy()
    new_win()

progress.place(x=-10, y=580)

############Button in Spalsh Screen##############

Frame(w, width=1024, height=580, bg=a).place(x=0, y=0)  # 249794
b1 = Button(w, width=20, height=3, text='Get Started',
            command=bar, border=0, fg=a, bg='white')
b1.place(x=430, y=500)


######## Label in Spalsh Screen ############

l1 = Label(w, text='Sistem Deteksi', fg='white', bg=a, anchor='center')
lst1 = ('Calibri (Body)', 32, 'bold')
l1.config(font=lst1)
l1.place(x=340, y=80)

l3 = Label(w, text='Penyakit Malaria',
           fg='white', bg=a, anchor='center')
lst3 = ('Calibri (Body)', 32, 'bold')
l3.config(font=lst3)
l3.place(x=320, y=130)

def kamera():
    imagefolder = "/home/pi/main-programqin/Foto/"
    cam = PiCamera()
#     camera.resolution = (2592, 1944)
    cam.rotation = 180 #90,180,270
    cam.framerate = 50
    cam.brightness = 53
    cam.rotation = 180 #90,180,270
    cam.start_preview()
    def on_press(key):
        if key== keyboard.Key.esc:
            cam.stop_preview()
            cam.close()
            return False
        elif key== keyboard.Key.enter:
            sum=0
            for i in range(1,100):
                filename = "/home/pi/main-programaqin/Foto/Malaria_Full/pasien"+str(i)+'.jpg'
                fileExist = os.path.isfile(filename)
    
                if(fileExist):
                    sum+=1
            
            sum=sum+1
            namafile='pasien'+format(sum)+'.jpg'
            cam.capture("/home/pi/main-programaqin/Foto/Malaria_Full/"+namafile)
            cam.stop_preview()
            cam.close()
            return False
        
    with keyboard.Listener(
        on_press=on_press) as listener:
        listener.join()

def openimage():
    filters = (("pic", "*.jpg *.png"),)
    path = tkinter.filedialog.askopenfilename(initialdir = "/home/pi/main-programaqin/Foto/Malaria_Full/",filetypes=filters,)
    if len(path) != 0:
        cv2.namedWindow("Foto Pasien")
        img = cv2 .imread(path)

######### Menampilkan gambar ##########
        input = cv2.resize(img, (880, 530))
                      
        cv2.imshow("Foto Pasien", input)
        cv2.moveWindow("Foto Pasien", 80, 20)
    #     def on_press(key):
    #         if key== keyboard.Key.esc:
        cv2.waitKey(5000)
        cv2.destroyAllWindows()
            
    #     with keyboard.Listener(
    #         on_press=on_press) as listener:
    #         listener.join()
    
def proses():
    #path = tkinter.filedialog.askopenfilename(initialdir = "/home/pi/main-programaqin/Foto/")
    filters = (("pic", "*.jpg *.png *.JPG *.PNG"),)
    path = tkinter.filedialog.askopenfilenames(initialdir = "/home/pi/main-programaqin/Foto/Malaria_Crop",filetypes=filters,)
    model_new = tf.keras.models.load_model("/home/pi/main-programaqin/modelvideoadam64100.h5")
    update = False
    parasit=0
    normal=0
    for file in path:
        print(file)
        update = True
        cv2.namedWindow("Original")
        img = cv2 .imread(file)

######### Menampilkan gambar ##########
        input = cv2.resize(img, (400, 400))
      
        text = ('Status : Terinfeksi')
        text1 = ('Status : Tidak Terinfeksi')
        letak = (10, 380)
        font = cv2.FONT_HERSHEY_SIMPLEX
        fontScale = 1
        thickness = 2
        warna = (30, 215, 28)
    
        imgtest = image.load_img(file, target_size=(64,64))
        xtesting=image.img_to_array(imgtest)
        xtesting=np.expand_dims(xtesting,axis=0)
        classes = model_new.predict(xtesting)
        if classes[0][0] == 1:
            print("Parasitized")
            cv2.putText(input, text, letak, font, fontScale,warna, thickness, cv2.LINE_AA)
            parasit+=1
        elif classes[0][1] == 1:
            print("Uninfected")
            normal+=1
            cv2.putText(input, text1, letak, font, fontScale,warna, thickness, cv2.LINE_AA)
        else :
            print("tidak ada")
                    
        cv2.imshow("Original", input)
        cv2.moveWindow("Original", 305, 75)
        cv2.waitKey(2000)
        cv2.destroyAllWindows()

####### Write to Xlsx file #######
    if update == True :
        wb = load_workbook('/home/pi/main-programaqin/Data/Report.xlsx')
        ws = wb.active
        sum=0
        for row in range (2,100):
            for col in range(1,2):
                char = get_column_letter(col)
                if(ws[char+str(row)].value!=None):
                    sum+=1

        nama="Data"+format(sum+1)
        ws['A'+str(sum+2)].value = sum+1
        if parasit>0 :
            ws['B'+str(sum+2)].value = "Malaria"
        elif parasit==0:
            ws['B'+str(sum+2)].value = "Aman"
        ws['C'+str(sum+2)].value = parasit
        ws['D'+str(sum+2)].value = normal
        wb.save('/home/pi/main-programaqin/Data/Report.xlsx')

def report():

    root = Tk()
    root.title('Get Report')
    root.geometry("500x500")
    root.config(bg ='#178a4f' )
    width_of_window = 500
    height_of_window = 300
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x_coordinate = (screen_width/2)-(width_of_window/2)
    y_coordinate = (screen_height/2)-(height_of_window/2)
    root.geometry("%dx%d+%d+%d" %(width_of_window,height_of_window,x_coordinate,y_coordinate))
    b1 = Button(root,width=20,height=3,text='Close',command=root.destroy,border=0,fg=a,bg='white')

    root.attributes('-fullscreen',True)

    # Create frame
    my_frame = Frame(root)
    my_frame.pack(pady=5)

    # Create treeview
    my_tree = ttk.Treeview(my_frame)

    # File open function
    def file_open():
        df = pd.read_excel("/home/pi/main-programaqin/Data/Report.xlsx")

        # Set up new treeview
        my_tree["column"] = list(df.columns)
        my_tree["show"] = "headings"
        # Loop thru column list for headers
        for column in my_tree["column"]:
            my_tree.heading(column, text=column)

        # Put data in treeview
        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            my_tree.insert("", "end", values=row)

        # Pack the treeview finally
        my_tree.pack()
    
    b1.pack(pady=5)

    file_open()
    root.mainloop()

def help():    
    help=Tk()
    help.title('Help')
    help.geometry('1024x600')
    help.config(bg ='#178a4f' )
    a='#178a4f'
    width_of_window = 1024
    height_of_window = 600
    screen_width = help.winfo_screenwidth()
    screen_height = help.winfo_screenheight()
    x_coordinate = (screen_width/2)-(width_of_window/2)
    y_coordinate = (screen_height/2)-(height_of_window/2)
    help.geometry("%dx%d+%d+%d" %(width_of_window,height_of_window,x_coordinate,y_coordinate))
    help.attributes('-fullscreen',True)

    l1 = Label(help, text = 'Help ',fg='white',bg=a, anchor = 'center' )
    lst1=('Calibri (Body)',32,'bold')
    l1.config(font=lst1)
    l1.place(x=440,y=10)

    l1 = Label(help, text = '1. Open Camera  ',fg='white',bg=a, anchor = 'center' )
    lst1=('Calibri (Body)',20,'bold')
    l1.config(font=lst1)
    l1.place(x=10,y=80)

    l1 = Label(help, text = 'Berfungsi untuk mengambil gambar melalui kamera',fg='white',bg=a, anchor = 'center' )
    lst1=('Calibri (Body)',16,'bold')
    l1.config(font=lst1)
    l1.place(x=10,y=120)

    l1 = Label(help, text = '2. Open Photo Pasien ',fg='white',bg=a, anchor = 'center' )
    lst1=('Calibri (Body)',20,'bold')
    l1.config(font=lst1)
    l1.place(x=10,y=180)

    l1 = Label(help, text = 'Berfungsi untuk melihat gambar yang sudah diambil dari pasien',fg='white',bg=a,anchor ='center')
    lst1=('Calibri (Body)',16,'bold')
    l1.config(font=lst1)
    l1.place(x=10,y=220)

    l1 = Label(help, text = '3. Process Image  ',fg='white',bg=a, anchor = 'center' )
    lst1=('Calibri (Body)',20,'bold')
    l1.config(font=lst1)
    l1.place(x=10,y=280)

    l1 = Label(help, text ='Berfungsi untuk menjalankan program pendeteksian malaria',fg='white',bg=a,
               anchor = 'center' )
    lst1=('Calibri (Body)',16,'bold')
    l1.config(font=lst1)
    l1.place(x=10,y=320)

    l1 = Label(help, text = '4. Exit Application  ',fg='white',bg=a, anchor = 'center' )
    lst1=('Calibri (Body)',20,'bold')
    l1.config(font=lst1)
    l1.place(x=10,y=380)

    l1 = Label(help, text = 'Berfungsi untuk keluar dari Aplikasi',fg='white',bg=a,
               anchor = 'center' )
    lst1=('Calibri (Body)',16,'bold')
    l1.config(font=lst1)
    l1.place(x=10,y=420)

    l1 = Label(help, text = '5. Get Report  ',fg='white',bg=a, anchor = 'center' )
    lst1=('Calibri (Body)',20,'bold')
    l1.config(font=lst1)
    l1.place(x=10,y=480)

    l1 = Label(help, text = 'Berfungsi untuk menampilkan data hasil perhitungan',fg='white',
               bg=a, anchor = 'center' )
    lst1=('Calibri (Body)',16,'bold')
    l1.config(font=lst1)
    l1.place(x=10,y=520)

    b1 = Button(help,width=20,height=3,text='Back',command=help.destroy,border=0,fg=a,bg='white')
    b1.place(x=860,y=530)

    help.mainloop()

w.mainloop()